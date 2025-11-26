"""
Unmerge cells and fill blank cells in previously merged regions
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Tuple
import numpy as np


class UnmergeProcessor:
    """Handles unmerging cells and filling blank cells"""
    
    def __init__(self, file_path: str):
        """
        Initialize processor
        
        Args:
            file_path: Path to Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def process_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Process a sheet: unmerge cells and fill blanks
        
        Args:
            sheet_name: Name of the sheet to process
        
        Returns:
            DataFrame with unmerged and filled cells
        """
        # Load workbook
        self.workbook = load_workbook(self.file_path, data_only=True)
        self.worksheet = self.workbook[sheet_name]
        
        # Step 1: Store merged cell information before unmerging
        merged_ranges = list(self.worksheet.merged_cells.ranges)
        merged_data = []
        
        for merged_range in merged_ranges:
            # Skip if range is invalid (None values)
            if (merged_range.min_row is None or merged_range.max_row is None or 
                merged_range.min_col is None or merged_range.max_col is None):
                continue
            
            try:
                # Get the top-left cell value (the one that contains the actual value)
                top_left_cell = self.worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                value = top_left_cell.value
                
                # Store range and value
                merged_data.append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col,
                    'value': value
                })
            except Exception as e:
                # Skip this merged range if we can't access it
                print(f"    Warning: Could not process merged range: {e}")
                continue
        
        # Step 2: Unmerge all cells (safely)
        try:
            # Get all merged cell ranges as strings first
            merged_cell_ranges = [str(mr) for mr in self.worksheet.merged_cells.ranges]
            for range_string in merged_cell_ranges:
                try:
                    self.worksheet.unmerge_cells(range_string)
                except Exception as e:
                    print(f"    Warning: Could not unmerge range {range_string}: {e}")
        except Exception as e:
            print(f"    Warning: Error during unmerge operation: {e}")
        
        # Step 3: Fill blank cells in previously merged regions
        for merge_info in merged_data:
            value = merge_info['value']
            if value is not None:
                # Fill all cells in the merged range with the value
                for row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                    for col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                        cell = self.worksheet.cell(row=row, column=col)
                        if cell.value is None or cell.value == '':
                            cell.value = value
        
        # Step 4: Convert to DataFrame
        # Get all data from the worksheet
        data = []
        for row in self.worksheet.iter_rows(values_only=True):
            data.append(row)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Close workbook
        self.workbook.close()
        
        return df
    
    def get_sample_rows(self, df: pd.DataFrame, n: int = 10) -> pd.DataFrame:
        """
        Get top N rows for header analysis
        
        Args:
            df: DataFrame
            n: Number of rows to sample
        
        Returns:
            First n rows of DataFrame
        """
        return df.head(n).copy()

